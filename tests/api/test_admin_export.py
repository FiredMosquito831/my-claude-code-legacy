"""Admin export endpoint tests: formats, fields, categorization, periods, full-DB.

Covers GET /admin/api/export for both the request-log and web-search scopes.
A seeded store with more rows than the 500-row page cap proves the export
bypasses the cap and streams the entire matching set.
"""

import csv
import io
import time
from datetime import UTC, datetime, timedelta

import pytest
from fastapi.testclient import TestClient

from my_claude_code.core.request_log import (
    RequestRecord,
    RouteAttempt,
    RouteAttemptOutcome,
    get_request_log_store,
)
from tests.api.support import create_test_app


@pytest.fixture
def client():
    return TestClient(create_test_app(), client=("127.0.0.1", 50000))


@pytest.fixture
def seeded_store(tmp_path):
    store = get_request_log_store(tmp_path / "requests.db")
    assert store is not None
    base = time.time() - 100
    # 12 rows, more than enough to exercise paging and the 500-row bypass.
    for index in range(12):
        store.enqueue(
            RequestRecord(
                id=f"r{index}",
                endpoint="/v1/messages" if index % 2 == 0 else "/v1/responses",
                protocol="anthropic",
                provider="p1" if index % 2 == 0 else "p2",
                resolved_model="m1" if index % 3 else "m2",
                ts_epoch=base + index,
                status="error" if index == 4 else "success",
                error_message="boom" if index == 4 else None,
                tokens_in=10 * index,
                tokens_out=index,
                cache_read_tokens=5 if index % 2 else None,
                cache_write_tokens=2 if index % 2 else None,
                tool_call_count=index if index % 2 else 0,
                thinking_chars=100 if index % 2 else 0,
                duration_ms=float(100 * (index + 1)),
                input_text="in" * 3000,
                output_text="out",
            )
        )
    store.close()
    yield store


def _export(client, **params):
    return client.get("/admin/api/export", params=params)


class TestLoopbackGuard:
    def test_remote_client_is_forbidden(self, client, seeded_store) -> None:
        remote = TestClient(client.app, client=("203.0.113.10", 50000))
        response = remote.get(
            "/admin/api/export",
            params={"format": "json", "scope": "requests"},
        )
        assert response.status_code == 403


class TestFormatCoverage:
    @pytest.mark.parametrize(
        ("fmt", "content_type", "extension"),
        (
            ("json", "application/json", "json"),
            ("csv", "text/csv", "csv"),
            (
                "xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "xlsx",
            ),
            ("txt", "text/plain", "txt"),
        ),
    )
    def test_format_streams_correct_type(
        self, client, seeded_store, fmt, content_type, extension
    ) -> None:
        response = _export(client, format=fmt, scope="requests")
        assert response.status_code == 200
        assert response.headers["content-type"].startswith(content_type)
        assert response.headers["content-disposition"].startswith("attachment")
        assert f".{extension}" in response.headers["content-disposition"]

    def test_json_returns_all_rows(self, client, seeded_store) -> None:
        response = _export(client, format="json", scope="requests")
        rows = response.json()
        assert len(rows) == 12
        # Default fields exclude bodies; id/ts_iso always present.
        assert all("id" in row and "ts_iso" in row for row in rows)

    def test_csv_has_bom_and_header(self, client, seeded_store) -> None:
        response = _export(client, format="csv", scope="requests")
        assert response.content.startswith(b"\xef\xbb\xbf")
        text = response.content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text))
        rows = list(reader)
        headers = rows[0]
        # Keyset/always columns lead the header; ID and Status are present.
        assert "ID" in headers
        assert "Status" in headers
        assert len(rows) == 13  # header + 12 rows
        # Data cells are present and aligned to the header (regression: the
        # renderer used label keys for lookup and produced empty data rows).
        provider_col = headers.index("Provider")
        data_row = rows[1]
        assert data_row[provider_col] in {"p1", "p2"}

    def test_csv_detail_has_real_data(self, client, seeded_store) -> None:
        response = _export(
            client,
            format="csv",
            scope="requests",
            fields="providers,models,tokens_out",
        )
        rows = list(csv.reader(io.StringIO(response.content.decode("utf-8-sig"))))
        headers = rows[0]
        provider_col = headers.index("Provider")
        model_col = headers.index("Resolved model")
        tokens_col = headers.index("Tokens out")
        assert len(rows) == 13  # header + 12 rows
        assert any(row[provider_col] == "p1" for row in rows[1:])
        assert any(row[model_col] in {"m1", "m2"} for row in rows[1:])
        assert any(row[tokens_col] and row[tokens_col] != "0" for row in rows[1:])

    def test_txt_is_readable_report(self, client, seeded_store) -> None:
        response = _export(client, format="txt", scope="requests")
        text = response.content.decode("utf-8")
        assert "Export" in text
        assert "ID" in text
        assert "Status" in text

    def test_txt_does_not_truncate_long_values(self, client, tmp_path) -> None:
        # A long model id must survive the report (regression: values were cut
        # to the header width, e.g. "nvidia/nemotro").
        long_model = "nvidia/nemotron-3-ultra-550b-a55b"
        store = get_request_log_store(tmp_path / "requests.db")
        assert store is not None
        store.enqueue(
            RequestRecord(
                id="long",
                endpoint="/v1/messages",
                protocol="anthropic",
                provider="nvidia_nim",
                resolved_model=long_model,
                requested_model=long_model,
                ts_epoch=time.time(),
                status="success",
                tokens_in=10,
                tokens_out=1,
            )
        )
        store.close()
        response = _export(
            client,
            format="txt",
            scope="requests",
            fields="providers,models",
        )
        text = response.content.decode("utf-8")
        assert long_model in text

    def test_xlsx_reads_back(self, client, seeded_store) -> None:
        pytest.importorskip("openpyxl")
        import io as _io

        import openpyxl

        response = _export(client, format="xlsx", scope="requests")
        workbook = openpyxl.load_workbook(_io.BytesIO(response.content))
        sheet = workbook.active
        assert sheet.max_row == 13  # header + 12 rows
        header_row = [
            sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)
        ]
        assert "ID" in header_row
        assert "Status" in header_row

    def test_xlsx_detail_has_real_data(self, client, seeded_store) -> None:
        pytest.importorskip("openpyxl")
        import io as _io

        import openpyxl

        response = _export(
            client,
            format="xlsx",
            scope="requests",
            fields="providers,models,tokens_out",
        )
        workbook = openpyxl.load_workbook(_io.BytesIO(response.content))
        sheet = workbook.active
        header_row = [
            sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)
        ]
        # openpyxl columns are 1-based; the header list is 0-based.
        provider_col = header_row.index("Provider") + 1
        model_col = header_row.index("Resolved model") + 1
        # Data cells must be populated (regression: empty cells).
        providers = {
            sheet.cell(row, provider_col).value for row in range(2, sheet.max_row + 1)
        }
        models = {
            sheet.cell(row, model_col).value for row in range(2, sheet.max_row + 1)
        }
        assert providers == {"p1", "p2"}
        assert models == {"m1", "m2"}


class TestFieldSelection:
    def test_fields_select_columns(self, client, seeded_store) -> None:
        response = _export(
            client,
            format="json",
            scope="requests",
            fields="providers,models,tokens_out",
        )
        rows = response.json()
        assert all("provider" in row for row in rows)
        assert all("resolved_model" in row for row in rows)
        assert all("tokens_out" in row for row in rows)
        # Deselected field columns are absent.
        assert all("thinking_chars" not in row for row in rows)

    def test_body_fields_are_uncapped(self, client, seeded_store) -> None:
        response = _export(
            client,
            format="json",
            scope="requests",
            fields="providers,input",
        )
        rows = response.json()
        assert any(len(row["input_text"]) > 4096 for row in rows)  # 3000 chars * 2

    def test_bad_field_returns_400(self, client, seeded_store) -> None:
        response = _export(client, format="json", scope="requests", fields="nope")
        assert response.status_code == 400


class TestCategorization:
    def test_grouped_order(self, client, seeded_store) -> None:
        response = _export(
            client,
            format="json",
            scope="requests",
            group_by="provider,period,model",
            fields="providers,models,error_rate,tokens_out",
        )
        rows = response.json()
        assert rows
        assert all(
            "provider" in row and "period" in row and "model" in row for row in rows
        )
        assert all(
            "requests" in row and "errors" in row and "error_rate" in row
            for row in rows
        )
        # Ordered by group dimensions: provider then period then model.
        keys = [(row["provider"], row["period"], row["model"]) for row in rows]
        assert keys == sorted(keys)

    def test_different_order_diffs(self, client, seeded_store) -> None:
        a = _export(
            client,
            format="json",
            scope="requests",
            group_by="provider,period",
            fields="providers,error_rate",
        ).json()
        b = _export(
            client,
            format="json",
            scope="requests",
            group_by="period,provider",
            fields="providers,error_rate",
        ).json()
        # Same data, different column ordering.
        assert [row["provider"] for row in a] == sorted(row["provider"] for row in a)
        assert all("provider" in row and "period" in row for row in b)

    def test_bad_group_dimension_returns_400(self, client, seeded_store) -> None:
        response = _export(client, format="json", scope="requests", group_by="bogus")
        assert response.status_code == 400


class TestPeriodWindow:
    def test_since_until_epoch_filters(self, client, seeded_store) -> None:
        # Seeded rows are timestamped now-100..now-89. A window covering half
        # of them must return only those rows, proving the filter applies.
        now = time.time()
        since = now - 105
        until = now - 95
        response = _export(
            client,
            format="json",
            scope="requests",
            since=str(since),
            until=str(until),
        )
        rows = response.json()
        assert 0 < len(rows) < 12
        for row in rows:
            assert since <= row["ts_epoch"] <= until

        # A window that misses everything returns an empty array.
        outside = _export(
            client,
            format="json",
            scope="requests",
            since=str(now - 10),
            until=str(now),
        ).json()
        assert outside == []

    def test_iso_bounds(self, client, seeded_store) -> None:
        since = (datetime.now(UTC) - timedelta(minutes=10)).isoformat()
        response = _export(client, format="json", scope="requests", since=since)
        assert response.status_code == 200

    def test_invalid_bound_returns_400(self, client, seeded_store) -> None:
        response = _export(client, format="json", scope="requests", since="garbage")
        assert response.status_code == 400

    def test_all_time_returns_everything(self, client, seeded_store) -> None:
        # No since/until = lifetime (the "All time" period sends no bounds).
        response = _export(client, format="json", scope="requests")
        assert len(response.json()) == 12


class TestMultiValueFilters:
    def test_multi_provider_returns_union(self, client, seeded_store) -> None:
        response = _export(
            client,
            format="json",
            scope="requests",
            provider="p1,p2",
        )
        rows = response.json()
        assert len(rows) == 12
        assert {row["provider"] for row in rows} == {"p1", "p2"}

    def test_multi_provider_subset(self, client, seeded_store) -> None:
        response = _export(client, format="json", scope="requests", provider="p1")
        rows = response.json()
        assert rows
        assert {row["provider"] for row in rows} == {"p1"}

    def test_multi_model_returns_union(self, client, seeded_store) -> None:
        response = _export(
            client,
            format="json",
            scope="requests",
            model="m1,m2",
        )
        rows = response.json()
        assert len(rows) == 12
        # Each row matched by resolved_model OR requested_model.
        assert {row["resolved_model"] for row in rows} == {"m1", "m2"}

    def test_multi_provider_with_since_combines(self, client, seeded_store) -> None:
        now = time.time()
        since = now - 95
        response = _export(
            client,
            format="json",
            scope="requests",
            provider="p1,p2",
            since=str(since),
        )
        rows = response.json()
        assert rows
        for row in rows:
            assert row["ts_epoch"] >= since


class TestFullDBBypass:
    def test_export_returns_more_than_500_rows(self, client, tmp_path) -> None:
        # The autouse _isolate_request_log fixture makes the app read
        # tmp_path/requests.db, so seed exactly that path.
        store = get_request_log_store(tmp_path / "requests.db")
        assert store is not None
        base = time.time() - 1000
        for index in range(600):
            store.enqueue(
                RequestRecord(
                    id=f"b{index}",
                    endpoint="/v1/messages",
                    protocol="anthropic",
                    provider="p1",
                    resolved_model="m1",
                    ts_epoch=base + index,
                    status="success",
                    tokens_in=1,
                    tokens_out=1,
                )
            )
        store.close()
        response = _export(client, format="json", scope="requests")
        assert len(response.json()) == 600


class TestDisabledStore:
    def test_disabled_returns_enabled_false(self, monkeypatch, tmp_path) -> None:
        monkeypatch.setenv("HOME", str(tmp_path))
        monkeypatch.setenv("USERPROFILE", str(tmp_path))
        monkeypatch.setenv("REQUEST_LOG_ENABLED", "false")
        app = create_test_app()
        client = TestClient(app, client=("127.0.0.1", 50000))
        response = _export(client, format="json", scope="requests")
        assert response.status_code == 200
        assert response.json() == {"enabled": False}


class TestWebSearchScope:
    def test_websearch_detail(self, client, monkeypatch, tmp_path) -> None:
        from my_claude_code.api.admin_websearch_routes import get_websearch_log_store
        from my_claude_code.websearch.analytics import WebSearchLogStore
        from my_claude_code.websearch.registry import SearchOutcome

        store = WebSearchLogStore(tmp_path / "ws.db")
        store.record(
            SearchOutcome(
                ts_epoch=time.time(),
                ts_iso=datetime.now(UTC).isoformat(),
                provider="exa",
                key_index=0,
                key_label="exak…1234",
                query="apple pie",
                results_count=5,
                duration_ms=12.5,
                status="success",
                error_kind=None,
                error_message=None,
                cost_usd=0.01,
                input_payload={"q": "apple pie"},
                output_payload={"answer": "a"},
                provider_config={"provider_id": "exa"},
            )
        )
        store.flush()
        monkeypatch.setenv("HOME", str(tmp_path))
        monkeypatch.setenv("USERPROFILE", str(tmp_path))
        app = create_test_app()
        app.dependency_overrides[get_websearch_log_store] = lambda: store
        ws_client = TestClient(app, client=("127.0.0.1", 50000))
        response = _export(
            ws_client,
            format="json",
            scope="websearch",
            fields="provider,results_count,cost_usd",
        )
        assert response.status_code == 200
        rows = response.json()
        assert len(rows) == 1
        assert rows[0]["provider"] == "exa"
        assert rows[0]["results_count"] == 5
        assert rows[0]["cost_usd"] == 0.01
        app.dependency_overrides.clear()
        store.close()

    def test_websearch_detail_export_carries_full_output(
        self, client, monkeypatch, tmp_path
    ) -> None:
        from my_claude_code.api.admin_websearch_routes import get_websearch_log_store
        from my_claude_code.websearch.analytics import WebSearchLogStore
        from my_claude_code.websearch.registry import SearchOutcome

        store = WebSearchLogStore(tmp_path / "ws.db")
        store.record(
            SearchOutcome(
                ts_epoch=time.time(),
                ts_iso=datetime.now(UTC).isoformat(),
                provider="exa",
                key_index=0,
                key_label="exak…1234",
                query="apple pie",
                results_count=1,
                duration_ms=12.5,
                status="success",
                error_kind=None,
                error_message=None,
                cost_usd=0.01,
                input_payload={"q": "apple pie", "max_results": 10},
                output_payload={
                    "answer": "a",
                    "results": [
                        {
                            "title": "Apple Pie",
                            "url": "https://example.com/pie",
                            "snippet": "A snippet.",
                            "content": "x" * 60_000,
                            "published": "",
                        }
                    ],
                },
                provider_config={"provider_id": "exa"},
            )
        )
        store.flush()
        monkeypatch.setenv("HOME", str(tmp_path))
        monkeypatch.setenv("USERPROFILE", str(tmp_path))
        app = create_test_app()
        app.dependency_overrides[get_websearch_log_store] = lambda: store
        ws_client = TestClient(app, client=("127.0.0.1", 50000))
        response = _export(
            ws_client,
            format="json",
            scope="websearch",
            fields="query,input,output,provider_config",
        )
        assert response.status_code == 200
        rows = response.json()
        assert len(rows) == 1
        # The decoded names must round-trip, not the raw *_json columns, and a
        # large output must arrive in full (regression: websearch export used
        # column names that are not DB columns and silently returned empty).
        assert rows[0]["query"] == "apple pie"
        assert rows[0]["input"]["q"] == "apple pie"
        assert rows[0]["output"]["answer"] == "a"
        assert rows[0]["output"]["results"][0]["content"] == "x" * 60_000
        assert rows[0]["provider_config"]["provider_id"] == "exa"
        app.dependency_overrides.clear()
        store.close()

    def test_websearch_multi_provider(self, client, monkeypatch, tmp_path) -> None:
        from my_claude_code.api.admin_websearch_routes import get_websearch_log_store
        from my_claude_code.websearch.analytics import WebSearchLogStore
        from my_claude_code.websearch.registry import SearchOutcome

        store = WebSearchLogStore(tmp_path / "ws.db")
        for index, provider in enumerate(("exa", "brave", "exa")):
            store.record(
                SearchOutcome(
                    ts_epoch=time.time() - index,
                    ts_iso=datetime.now(UTC).isoformat(),
                    provider=provider,
                    key_index=0,
                    key_label=f"{provider[:3]}…abcd",
                    query="q",
                    results_count=3,
                    duration_ms=10.0,
                    status="success",
                    error_kind=None,
                    error_message=None,
                    cost_usd=0.01,
                    input_payload={"q": "q"},
                    output_payload={"answer": "a"},
                    provider_config={"provider_id": provider},
                )
            )
        store.flush()
        monkeypatch.setenv("HOME", str(tmp_path))
        monkeypatch.setenv("USERPROFILE", str(tmp_path))
        app = create_test_app()
        app.dependency_overrides[get_websearch_log_store] = lambda: store
        ws_client = TestClient(app, client=("127.0.0.1", 50000))
        response = _export(
            ws_client,
            format="json",
            scope="websearch",
            fields="provider",
            provider="exa,brave",
        )
        assert response.status_code == 200
        rows = response.json()
        assert len(rows) == 3
        assert {row["provider"] for row in rows} == {"exa", "brave"}
        app.dependency_overrides.clear()
        store.close()


class TestLadderColumns:
    """``request_attempts`` is joined by no export SQL; these are the bridge."""

    @pytest.fixture
    def ladder_store(self, tmp_path):
        store = get_request_log_store(tmp_path / "requests.db")
        assert store is not None
        store.enqueue(
            RequestRecord(
                id="req_ladder",
                endpoint="/v1/messages",
                protocol="anthropic",
                provider="nvidia_nim",
                resolved_model="kimi-k3",
                ts_epoch=time.time(),
                status="error",
                error_message="Upstream provider NIM returned HTTP 502.",
                attempts=(
                    RouteAttempt(
                        attempt=0,
                        provider="nvidia_nim",
                        model_ref="nvidia_nim/kimi-k3",
                        outcome=RouteAttemptOutcome.FAILED,
                        error_kind="upstream",
                        error_message="HTTP 502",
                        duration_ms=107_534.16,
                        ladder_tries=3,
                        params={
                            "ladder": {
                                "tries": [
                                    {"source": "upstream", "status": 429},
                                    {"source": "upstream", "status": 429},
                                    {"source": "upstream", "status": 502},
                                ],
                                "summary": {
                                    "tries": 3,
                                    "statuses_by_code": {"429": 2, "502": 1},
                                    "keys": 1,
                                    "time_sleeping_ms": 6000.0,
                                    "time_limiter_ms": 0.0,
                                    "tries_dropped": 0,
                                },
                                "credentials": [],
                                "root_cause": "1 keys \N{MULTIPLICATION SIGN} 3 tries: 2\N{MULTIPLICATION SIGN}429, 1\N{MULTIPLICATION SIGN}502",
                            }
                        },
                    ),
                ),
            )
        )
        store.close()
        yield store

    def test_ladder_columns_are_declared_derived(self) -> None:
        """Guards ``tests/api/test_reasoning_recording.py``'s schema assertion.

        They come from a join, not from a ``requests`` column, so naming them
        to the SELECT would break every detail export.
        """
        from my_claude_code.core import export as export_engine

        assert export_engine.request_detail_columns(["ladder"]) == (
            export_engine.request_detail_columns([])
        )
        assert export_engine.request_detail_derived_columns(["ladder"]) == [
            "ladder_tries",
            "ladder_statuses",
            "ladder_root_cause",
        ]

    def test_ladder_columns_are_selectable_and_reach_all_four_formats(
        self, client, ladder_store
    ) -> None:
        expected = (
            "3",
            "429\N{MULTIPLICATION SIGN}2, 502\N{MULTIPLICATION SIGN}1",
            "1 keys \N{MULTIPLICATION SIGN} 3 tries: 2\N{MULTIPLICATION SIGN}429, 1\N{MULTIPLICATION SIGN}502",
        )

        rows = _export(client, format="json", scope="requests", fields="ladder").json()
        assert rows[0]["ladder_tries"] == 3
        assert rows[0]["ladder_statuses"] == expected[1]
        assert rows[0]["ladder_root_cause"] == expected[2]

        csv_text = _export(
            client, format="csv", scope="requests", fields="ladder"
        ).content.decode("utf-8-sig")
        assert "Upstream tries" in csv_text
        assert expected[1] in csv_text
        assert expected[2] in csv_text

        txt = _export(
            client, format="txt", scope="requests", fields="ladder"
        ).content.decode("utf-8")
        assert expected[1] in txt
        assert expected[2] in txt

        pytest.importorskip("openpyxl")
        import io as _io

        import openpyxl

        workbook = openpyxl.load_workbook(
            _io.BytesIO(
                _export(
                    client, format="xlsx", scope="requests", fields="ladder"
                ).content
            )
        )
        sheet = workbook.active
        header = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
        assert "Upstream statuses" in header
        cell = sheet.cell(2, header.index("Upstream statuses") + 1).value
        assert cell == expected[1]

    def test_ladder_columns_are_absent_when_the_field_is_not_selected(
        self, client, ladder_store
    ) -> None:
        rows = _export(client, format="json", scope="requests", fields="models").json()
        assert all("ladder_tries" not in row for row in rows)
